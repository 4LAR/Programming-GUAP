import openpyxl

# Создание Excel-файла
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Boundary Value Problem Solution"

# Запись заголовков
ws.append(["x", "y(x)", "y'(x)"])

# Заполнение значений x и формул для y(x) и y'(x)
num_points = 100
x_start = 1
x_end = 4
dx = (x_end - x_start) / (num_points - 1)

# Заполнение значений x
for i in range(num_points):
    x_value = x_start + i * dx
    ws.cell(row=i + 2, column=1, value=x_value)  # Запись x

    # Формулы для y(x) и y'(x)
    if i == 0:
        ws.cell(row=i + 2, column=2, value=2)  # Начальное значение y(1) = 2
        ws.cell(row=i + 2, column=3, value="= (3/4) * B2 / A2 - (1/2) * (A2^(-1/2)) * C2")  # y'(1)
    elif i == num_points - 1:
        ws.cell(row=i + 2, column=2, value="=4.5")  # Конечное значение y(4) = 4.5
        ws.cell(row=i + 2, column=3, value="= (3/4) * B" + str(i + 2) + " / A" + str(i + 2) + " - (1/2) * (A" + str(i + 2) + "^(-1/2)) * C" + str(i + 1))
    else:
        ws.cell(row=i + 2, column=2, value="=B" + str(i + 1) + " + D" + str(i + 2))  # y(x) = y(x-1) + dy
        ws.cell(row=i + 2, column=3, value="= (3/4) * B" + str(i + 2) + " / A" + str(i + 2) + " - (1/2) * (A" + str(i + 2) + "^(-1/2)) * C" + str(i + 1))

    # Запись шага
    if i > 0:
        ws.cell(row=i + 2, column=4, value="=B" + str(i + 1) + " - B" + str(i))  # dy

# Сохранение файла
wb.save("bvp_solution_with_formulas.xlsx")

print("Excel файл 'bvp_solution_with_formulas.xlsx' успешно создан.")
